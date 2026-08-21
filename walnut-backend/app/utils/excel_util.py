"""Excel 导入导出工具（openpyxl 实现）。

- 导出：``export_excel_response(rows, header_mapping, sheet_name)`` 返回带
  ``Content-Disposition`` 的流式响应（文件名 UTF-8 编码，兼容前端下载组件）；
- 导入：``read_excel_to_dicts(contents)`` 将上传的 Excel 字节解析为字典列表（首行表头）。
"""

import io
from datetime import date, datetime, time
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.common.constant import DATE_DISPLAY_FMT, DATETIME_DISPLAY_FMT
from app.common.response import StreamResponse


def _fmt_cell(value: Any) -> Any:
    """单元格值格式化：日期时间转展示字符串，None 转空串。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(DATETIME_DISPLAY_FMT)
    if isinstance(value, date):
        return value.strftime(DATE_DISPLAY_FMT)
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


class ExcelUtil:
    """Excel 相关处理。"""

    EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @staticmethod
    def read_excel_to_dicts(contents: bytes) -> list[dict[str, Any]]:
        """读取 Excel 文件字节，返回字典列表（首行为列名）。"""
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise ValueError("工作簿没有活动工作表")
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        result: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_dict: dict[str, Any] = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_dict[str(headers[i])] = val
            if row_dict:
                result.append(row_dict)
        wb.close()
        return result

    @staticmethod
    def build_workbook(rows: list[dict[str, Any]], header_mapping: dict[str, str], sheet_name: str) -> bytes:
        """按 ``header_mapping``（{字段名: 中文表头}，顺序即列序）生成工作簿字节。"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        header_fill = PatternFill(start_color="ababab", end_color="ababab", fill_type="solid")
        headers = list(header_mapping.values())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 16

        for row_num, item in enumerate(rows, 2):
            for col_num, field in enumerate(header_mapping.keys(), 1):
                ws.cell(row=row_num, column=col_num).value = _fmt_cell(item.get(field))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def export_excel_response(
        cls,
        rows: list[dict[str, Any]],
        header_mapping: dict[str, str],
        sheet_name: str,
        filename: str | None = None,
    ) -> StreamResponse:
        """导出 Excel 并返回下载响应。"""
        from app.utils.date_util import date_time_now

        data = cls.build_workbook(rows, header_mapping, sheet_name)
        filename = filename or f"{sheet_name}_{date_time_now()}.xlsx"
        encoded = quote(filename)
        headers = {
            "Content-Disposition": f"attachment;filename*=utf-8''{encoded}",
            "download-filename": encoded,
            "Access-Control-Expose-Headers": "Content-Disposition,download-filename",
        }
        return StreamResponse(
            data=iter([data]),
            media_type=cls.EXCEL_CONTENT_TYPE,
            headers=headers,
        )

    @staticmethod
    def get_excel_template(header_list: list[str], selector_header_list: list[str], option_list: list[dict[str, list[str]]]) -> bytes:
        """生成带下拉选项的 Excel 导入模板。"""
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        header_fill = PatternFill(start_color="ababab", end_color="ababab", fill_type="solid")

        for col_num, header in enumerate(header_list, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 14

        for selector_header in selector_header_list:
            if selector_header not in header_list:
                continue
            col_idx = header_list.index(selector_header) + 1
            options = next((item.get(selector_header, []) for item in option_list if selector_header in item), [])
            if not options:
                continue
            formula = '"' + ",".join(str(o) for o in options) + '"'
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}1001")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
